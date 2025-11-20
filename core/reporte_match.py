# core/reporte_match.py
# -- coding: utf-8 --
"""
Genera el reporte de validación final (ENVIAR / CUADRAR)
desde comparacion_match.
"""

import os
import sqlite3
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

DB_PATH = os.path.join(os.getcwd(), "database", "integrador.db")
OUTPUT_FILE = os.path.join(os.getcwd(), "reportes", f"REPORTE_MATCH_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

def _conn():
    return sqlite3.connect(DB_PATH)

def map_transaction_kind(kind):
    if not isinstance(kind, str): return kind
    k = kind.strip()
    if k == "PICHINCHA CTA 3474862904": return "91qdGwQNiLvEdN8j"
    if k == "COOP. DE AHORRO Y CREDITO PEDRO MONCAYO LTDA. CTA 241701040196": return "gDGe7DYVFWD2an2x"
    if k == "PROCREDIT S.A. CTA 019037892134": return "1mBdJ14VsQNlb0J6"
    return k

def generar_reporte_match():
    con = _conn()
    df = pd.read_sql_query("""
        SELECT documento_id, total_contifico, total_wispro,
               documento_numero, transaction_kind, transaction_code,
               fecha_emision, created_at AS created_at_texto, estado_final
        FROM comparacion_match
    """, con)
    con.close()

    if df.empty:
        print("⚠ No hay registros en comparacion_match.")
        return {"match": 0}

    df["codigo_mapeado"] = df["transaction_kind"].apply(map_transaction_kind)
    df["B_forma_cobro"] = "TRA"

    # Validación: verde o azul
    df["validacion"] = df["estado_final"].apply(
        lambda x: "ENVIAR" if str(x).upper() == "APROBADO" else "CUADRAR"
    )

    df.to_excel(OUTPUT_FILE, index=False)
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active

    # Colorear columna validacion
    headers = [cell.value for cell in ws[1]]
    col_val = headers.index("validacion") + 1
    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    azul = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    verde_font = Font(color="006100", bold=True)
    azul_font = Font(color="1F4E78", bold=True)

    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=col_val)
        val = str(cell.value).strip().upper()
        if val == "ENVIAR":
            cell.fill = verde; cell.font = verde_font
        else:
            cell.fill = azul; cell.font = azul_font

    for c in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = 22

    wb.save(OUTPUT_FILE)
    return {"match": len(df), "archivo": OUTPUT_FILE}

def listar_para_tabla():
    con = _conn()
    df = pd.read_sql_query("""
        SELECT documento_id, total_contifico, total_wispro,
               documento_numero, transaction_kind, transaction_code,
               fecha_emision, created_at AS created_at_texto, estado_final
        FROM comparacion_match
        ORDER BY creado_en DESC
    """, con)
    con.close()
    if df.empty: return []
    df["codigo_mapeado"] = df["transaction_kind"].apply(map_transaction_kind)
    df["validacion"] = df["estado_final"].apply(
        lambda x: "ENVIAR" if str(x).upper() == "APROBADO" else "CUADRAR"
    )
    return df.to_dict(orient="records")
